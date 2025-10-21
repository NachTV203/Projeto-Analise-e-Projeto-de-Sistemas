import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import ttkbootstrap as tb
from ttkbootstrap.constants import *
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import os
import sqlite3
import re
# Importações necessárias do openpyxl para formatação avançada
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Constantes ---
DB_FILE = 'financeiro.db'
RELATORIOS_DIR = 'Relatorios_Eventos'


# --- CLASSE: GERENCIADOR DE EVENTOS (SEM MUDANÇAS NA LÓGICA, APENAS CHAMA A NOVA EXPORTAÇÃO) ---
class EventosManagerWindow(tb.Toplevel):
    def __init__(self, parent, app_controller):
        super().__init__(parent);
        self.app_controller = app_controller;
        self.title("Gerenciador de Eventos");
        self.geometry("700x500")
        self.criar_widgets();
        self.carregar_eventos_tree();
        self.atualizar_botoes_acao();
        self.transient(parent);
        self.grab_set()

    def criar_novo_evento(self):
        novo_evento = self.entry_novo_evento.get().strip()
        if not novo_evento: messagebox.showwarning("Atenção", "Nome do evento não pode ser vazio.", parent=self); return
        eventos_existentes = self.app_controller.df_dividendos['Evento'].dropna().unique().tolist()
        if novo_evento in eventos_existentes: messagebox.showwarning("Atenção", "Este evento já existe.",
                                                                     parent=self); return
        nova_linha = pd.DataFrame(
            [{'Evento': novo_evento, 'Credor': None, 'Descricao': None, 'Valor': None, 'Status': None}])
        self.app_controller.df_dividendos = pd.concat([self.app_controller.df_dividendos, nova_linha],
                                                      ignore_index=True)
        self.app_controller.salvar_dados()
        self.app_controller.exportar_evento_para_excel(novo_evento)
        self.carregar_eventos_tree();
        self.entry_novo_evento.delete(0, END)
        messagebox.showinfo("Sucesso", f"Evento '{novo_evento}' criado e relatório Excel gerado.", parent=self)

    def editar_evento_selecionado(self):
        nome_antigo = self.get_evento_selecionado();
        if not nome_antigo: return
        nome_novo = simpledialog.askstring("Editar Evento", f"Novo nome para '{nome_antigo}':", parent=self)
        if not nome_novo or not nome_novo.strip(): return
        nome_novo = nome_novo.strip();
        eventos_existentes = self.app_controller.df_dividendos['Evento'].unique().tolist()
        if nome_novo in eventos_existentes and nome_novo != nome_antigo: messagebox.showerror("Erro",
                                                                                              "Nome de evento já existe.",
                                                                                              parent=self); return
        df = self.app_controller.df_dividendos;
        df.loc[df['Evento'] == nome_antigo, 'Evento'] = nome_novo
        self.app_controller.salvar_dados()
        self.app_controller.renomear_relatorio_excel(nome_antigo, nome_novo)
        self.app_controller.exportar_evento_para_excel(nome_novo)
        self.carregar_eventos_tree()

    def excluir_evento_selecionado(self):
        nome_evento = self.get_evento_selecionado();
        if not nome_evento: return
        if messagebox.askyesno("Confirmar Exclusão",
                               f"Excluir o evento '{nome_evento}', suas dívidas E seu relatório Excel?", parent=self):
            df = self.app_controller.df_dividendos;
            self.app_controller.df_dividendos = df[df['Evento'] != nome_evento]
            self.app_controller.salvar_dados()
            self.app_controller.excluir_relatorio_excel(nome_evento)
            self.carregar_eventos_tree();
            self.atualizar_botoes_acao()

    def criar_widgets(self):
        frame = tb.Frame(self, padding=20);
        frame.pack(expand=True, fill=BOTH)
        create_frame = tb.Labelframe(frame, text=" 1. Criar Novo Evento ", bootstyle=INFO);
        create_frame.pack(fill=X, pady=10)
        tb.Label(create_frame, text="Nome do Evento:").pack(side=LEFT, padx=10, pady=10)
        self.entry_novo_evento = tb.Entry(create_frame, bootstyle=PRIMARY);
        self.entry_novo_evento.pack(side=LEFT, fill=X, expand=True, padx=(0, 10), pady=10)
        tb.Button(create_frame, text="Criar Evento", bootstyle=SUCCESS, command=self.criar_novo_evento).pack(side=RIGHT,
                                                                                                             padx=10,
                                                                                                             pady=10)
        list_frame = tb.Labelframe(frame, text=" 2. Selecione um Evento na Lista ", bootstyle=INFO);
        list_frame.pack(fill=BOTH, expand=True, pady=10)
        self.tree_eventos = tb.Treeview(list_frame, columns=("Evento",), show='headings', bootstyle=PRIMARY);
        self.tree_eventos.heading("Evento", text="Nome do Evento");
        self.tree_eventos.pack(fill=BOTH, expand=True)
        self.tree_eventos.bind("<<TreeviewSelect>>", self.atualizar_botoes_acao)
        action_frame = tb.Labelframe(frame, text=" 3. Ações para o Evento Selecionado ", bootstyle=INFO);
        action_frame.pack(fill=X, pady=10)
        self.btn_abrir = tb.Button(action_frame, text="Abrir / Ver Dívidas", bootstyle=PRIMARY,
                                   command=self.abrir_evento_selecionado);
        self.btn_abrir.pack(side=LEFT, expand=True, fill=X, padx=5, pady=5)
        self.btn_editar = tb.Button(action_frame, text="Editar Nome", bootstyle=SECONDARY,
                                    command=self.editar_evento_selecionado);
        self.btn_editar.pack(side=LEFT, expand=True, fill=X, padx=5, pady=5)
        self.btn_excluir = tb.Button(action_frame, text="Excluir Evento", bootstyle=DANGER,
                                     command=self.excluir_evento_selecionado);
        self.btn_excluir.pack(side=LEFT, expand=True, fill=X, padx=5, pady=5)

    def carregar_eventos_tree(self):
        self.tree_eventos.delete(*self.tree_eventos.get_children())
        if 'Evento' in self.app_controller.df_dividendos.columns:
            eventos_validos = self.app_controller.df_dividendos['Evento'].dropna().unique().tolist()
            for evento in sorted(eventos_validos): self.tree_eventos.insert('', 'end', values=(evento,))

    def atualizar_botoes_acao(self, event=None):
        estado = NORMAL if self.tree_eventos.selection() else DISABLED
        self.btn_abrir.config(state=estado);
        self.btn_editar.config(state=estado);
        self.btn_excluir.config(state=estado)

    def get_evento_selecionado(self):
        selecionado = self.tree_eventos.focus();
        return self.tree_eventos.item(selecionado)['values'][0] if selecionado else None

    def abrir_evento_selecionado(self):
        nome_evento = self.get_evento_selecionado()
        if nome_evento: DividendosWindow(self, self.app_controller, nome_evento)


# --- CLASSE DIVIDENDOS (com a lógica de exportação) ---
class DividendosWindow(tb.Toplevel):
    def __init__(self, parent, app_controller, evento_selecionado):
        super().__init__(parent);
        self.app_controller = app_controller;
        self.evento_selecionado = evento_selecionado
        self.title(f"Dívidas do Evento: {self.evento_selecionado}");
        self.geometry("800x500")
        self.criar_widgets();
        self.carregar_dividendos_tree();
        self.transient(parent);
        self.grab_set()

    def adicionar_dividendo(self):
        credor = self.credor_entry.get();
        descricao = self.desc_entry_div.get();
        valor_str = self.valor_entry_div.get().replace(',', '.')
        if not all([credor, descricao, valor_str]): messagebox.showerror("Erro", "Campos obrigatórios.",
                                                                         parent=self); return
        try:
            valor = float(valor_str)
        except ValueError:
            messagebox.showerror("Erro", "Valor inválido.", parent=self);
            return
        df = self.app_controller.df_dividendos;
        condicao_fantasma = (df['Evento'] == self.evento_selecionado) & (df['Credor'].isnull());
        self.app_controller.df_dividendos = df.drop(df[condicao_fantasma].index)
        novo_dividendo = {'Evento': self.evento_selecionado, 'Credor': credor, 'Descricao': descricao, 'Valor': valor,
                          'Status': 'Em Aguardo'}
        self.app_controller.df_dividendos = pd.concat(
            [self.app_controller.df_dividendos, pd.DataFrame([novo_dividendo])], ignore_index=True)
        self.app_controller.salvar_dados();
        self.app_controller.exportar_evento_para_excel(self.evento_selecionado);
        self.carregar_dividendos_tree()
        self.credor_entry.delete(0, END);
        self.desc_entry_div.delete(0, END);
        self.valor_entry_div.delete(0, END)

    def alterar_status(self, novo_status):
        selecionado = self.tree_dividendos.selection()
        if not selecionado: return

        for item_id in selecionado:
            index = int(item_id)
            # Pega o status antigo para evitar registrar despesa duplicada
            status_antigo = self.app_controller.df_dividendos.at[index, 'Status']

            # Se o novo status for "Pago" E o antigo NÃO era "Pago", registra a despesa
            if novo_status == 'Pago' and status_antigo != 'Pago':
                divida = self.app_controller.df_dividendos.loc[index]
                self.app_controller.registrar_pagamento_divida(
                    nome_evento=self.evento_selecionado,
                    credor=divida['Credor'],
                    descricao_divida=divida['Descricao'],
                    valor_divida=divida['Valor']
                )

            # Atualiza o status no DataFrame de dividendos
            self.app_controller.df_dividendos.at[index, 'Status'] = novo_status

        # Salva os dados, exporta o relatório atualizado e recarrega a lista
        self.app_controller.salvar_dados()
        self.app_controller.exportar_evento_para_excel(self.evento_selecionado)
        self.carregar_dividendos_tree()

    def excluir_dividendo(self):
        selecionado = self.tree_dividendos.selection();
        if not selecionado or not messagebox.askyesno("Confirmar", "Tem certeza?", parent=self): return
        indices = [int(item_id) for item_id in selecionado]
        self.app_controller.df_dividendos = self.app_controller.df_dividendos.drop(indices).reset_index(drop=True)
        self.app_controller.salvar_dados();
        self.app_controller.exportar_evento_para_excel(self.evento_selecionado);
        self.carregar_dividendos_tree()

    def criar_widgets(self):
        frame = tb.Frame(self, padding=20);
        frame.pack(expand=True, fill=BOTH)
        form_frame = tb.Labelframe(frame, text=" Adicionar Nova Dívida/Dividendo ", bootstyle=INFO);
        form_frame.pack(fill=X, pady=10)
        tb.Label(form_frame, text="Credor/Devedor:").grid(row=0, column=0, padx=5, pady=5, sticky='w');
        self.credor_entry = tb.Entry(form_frame, bootstyle=PRIMARY);
        self.credor_entry.grid(row=0, column=1, padx=5, pady=5, sticky='ew')
        tb.Label(form_frame, text="Descrição:").grid(row=0, column=2, padx=5, pady=5, sticky='w');
        self.desc_entry_div = tb.Entry(form_frame, bootstyle=PRIMARY);
        self.desc_entry_div.grid(row=0, column=3, padx=5, pady=5, sticky='ew')
        tb.Label(form_frame, text="Valor (R$):").grid(row=1, column=0, padx=5, pady=5, sticky='w');
        self.valor_entry_div = tb.Entry(form_frame, bootstyle=PRIMARY);
        self.valor_entry_div.grid(row=1, column=1, padx=5, pady=5, sticky='ew')
        tb.Button(form_frame, text="Adicionar", bootstyle=SUCCESS, command=self.adicionar_dividendo).grid(row=1,
                                                                                                          column=3,
                                                                                                          padx=5,
                                                                                                          pady=5,
                                                                                                          sticky='ew')
        form_frame.columnconfigure(1, weight=1);
        form_frame.columnconfigure(3, weight=1)
        list_frame = tb.Labelframe(frame, text=" Lista de Dívidas/Dividendos ", bootstyle=INFO);
        list_frame.pack(fill=BOTH, expand=True, pady=10)
        cols_div = ('Credor', 'Descricao', 'Valor', 'Status');
        self.tree_dividendos = tb.Treeview(list_frame, columns=cols_div, show='headings', bootstyle=PRIMARY)
        for col in cols_div: self.tree_dividendos.heading(col, text=col)
        self.tree_dividendos.pack(fill=BOTH, expand=True, side=LEFT)
        self.tree_dividendos.tag_configure('pago', background='#d4edda', foreground='black');
        self.tree_dividendos.tag_configure('aguardando', background='#fff3cd', foreground='black')
        action_list_frame = tb.Frame(frame);
        action_list_frame.pack(fill=X, pady=5)
        tb.Button(action_list_frame, text="Marcar como Pago", bootstyle=(SUCCESS, OUTLINE),
                  command=lambda: self.alterar_status('Pago')).pack(side=LEFT, padx=5)
        tb.Button(action_list_frame, text="Marcar como 'Em Aguardo'", bootstyle=(WARNING, OUTLINE),
                  command=lambda: self.alterar_status('Em Aguardo')).pack(side=LEFT, padx=5)
        tb.Button(action_list_frame, text="Excluir", bootstyle=(DANGER, OUTLINE), command=self.excluir_dividendo).pack(
            side=LEFT, padx=5)

    def carregar_dividendos_tree(self):
        self.tree_dividendos.delete(*self.tree_dividendos.get_children())
        df_evento = self.app_controller.df_dividendos[
            self.app_controller.df_dividendos['Evento'] == self.evento_selecionado]
        df_filtrado = df_evento.dropna(subset=['Credor'])
        for index, row in df_filtrado.iterrows():
            tag = 'pago' if row['Status'] == 'Pago' else 'aguardando'
            valores_exibidos = [row['Credor'], row['Descricao'], row['Valor'], row['Status']]
            self.tree_dividendos.insert('', 'end', iid=index, values=valores_exibidos, tags=(tag,))


# --- CLASSE RELATÓRIO MENSAL (SEM MUDANÇAS) ---
class RelatorioMensalWindow(tb.Toplevel):
    def __init__(self, parent, app_controller):
        super().__init__(parent);
        self.app_controller = app_controller;
        self.title("Relatório Financeiro Anual");
        self.geometry("800x600")
        self.criar_widgets();
        self.popular_seletor_ano();
        self.transient(parent);
        self.grab_set()

    def criar_widgets(self):
        frame = tb.Frame(self, padding=20);
        frame.pack(expand=True, fill=BOTH)
        filtro_frame = tb.Frame(frame);
        filtro_frame.pack(fill=X, pady=(0, 10))
        tb.Label(filtro_frame, text="Selecione o Ano para Gerar o Relatório:", font=("Helvetica", 12)).pack(side=LEFT,
                                                                                                            padx=(
                                                                                                                0, 10))
        self.combo_ano_relatorio = tb.Combobox(filtro_frame, bootstyle=PRIMARY, state="readonly");
        self.combo_ano_relatorio.pack(side=LEFT, fill=X, expand=True)
        self.combo_ano_relatorio.bind("<<ComboboxSelected>>", self.gerar_e_exibir_relatorio)
        report_frame = tb.Labelframe(frame, text=" Resumo Mensal ", bootstyle=INFO);
        report_frame.pack(fill=BOTH, expand=True, pady=10)
        cols = ("Mês", "Receitas", "Despesas", "Saldo");
        self.tree_relatorio = tb.Treeview(report_frame, columns=cols, show="headings", bootstyle=PRIMARY)
        for col in cols: self.tree_relatorio.heading(col, text=col); self.tree_relatorio.column(col, anchor=CENTER)
        self.tree_relatorio.pack(fill=BOTH, expand=True)
        self.tree_relatorio.tag_configure('lucro', background='#d4edda', foreground='black');
        self.tree_relatorio.tag_configure('prejuizo', background='#f8d7da', foreground='black')

    def popular_seletor_ano(self):
        df = self.app_controller.df_transacoes
        if not df.empty:
            anos = sorted(df['Data'].dt.year.unique(), reverse=True)
            self.combo_ano_relatorio['values'] = anos
            if anos: self.combo_ano_relatorio.current(0); self.gerar_e_exibir_relatorio()

    def gerar_e_exibir_relatorio(self, event=None):
        self.tree_relatorio.delete(*self.tree_relatorio.get_children())
        try:
            ano_selecionado = int(self.combo_ano_relatorio.get())
        except (ValueError, TypeError):
            return
        df = self.app_controller.df_transacoes;
        df_ano = df[df['Data'].dt.year == ano_selecionado].copy()
        if df_ano.empty: return
        df_ano['Mes'] = df_ano['Data'].dt.month
        receitas = df_ano[df_ano['Tipo'] == 'Receita'].groupby('Mes')['Valor'].sum()
        despesas = df_ano[df_ano['Tipo'] == 'Despesa'].groupby('Mes')['Valor'].sum()
        relatorio = pd.DataFrame({'Receitas': receitas, 'Despesas': despesas}).fillna(0);
        relatorio['Saldo'] = relatorio['Receitas'] - relatorio['Despesas']
        todos_meses = pd.DataFrame(index=range(1, 13));
        relatorio = todos_meses.join(relatorio).fillna(0)
        for mes_num, row in relatorio.iterrows():
            nome_mes = self.app_controller.MESES_NOME[mes_num];
            receita_str = f"R$ {row['Receitas']:,.2f}";
            despesa_str = f"R$ {row['Despesas']:,.2f}";
            saldo_str = f"R$ {row['Saldo']:,.2f}"
            tag = 'lucro' if row['Saldo'] >= 0 else 'prejuizo'
            self.tree_relatorio.insert("", "end", values=(nome_mes, receita_str, despesa_str, saldo_str), tags=(tag,))


# --- CLASSE TRANSAÇÃO (SEM MUDANÇAS) ---
class TransacaoDialog(tb.Toplevel):
    def __init__(self, parent, tipo, callback):
        super().__init__(parent);
        self.tipo = tipo;
        self.callback = callback;
        self.title(f"Adicionar {self.tipo}");
        self.geometry("400x360")
        frame = tb.Frame(self, padding=20);
        frame.pack(expand=True, fill=BOTH)
        tb.Label(frame, text="Data da Transação:").pack(pady=(0, 5))
        self.date_entry = tb.DateEntry(frame, bootstyle=PRIMARY, dateformat='%Y-%m-%d', firstweekday=0);
        self.date_entry.pack(fill=X, pady=(0, 10))
        label_desc = "Categoria (Ex: Aluguel):" if self.tipo == 'Despesa' else "Descrição (Ex: Cliente X):"
        tb.Label(frame, text=label_desc).pack(pady=(10, 5))
        self.desc_entry = tb.Entry(frame, bootstyle=PRIMARY);
        self.desc_entry.pack(fill=X, pady=(0, 10))
        tb.Label(frame, text="Valor (R$):").pack(pady=(10, 5));
        self.valor_entry = tb.Entry(frame, bootstyle=PRIMARY);
        self.valor_entry.pack(fill=X, pady=(0, 10))
        tb.Button(frame, text="Salvar", command=self.salvar, bootstyle=SUCCESS).pack(pady=20, fill=X)
        self.transient(parent);
        self.grab_set()

    def salvar(self):
        data_transacao = self.date_entry.entry.get()
        descricao = self.desc_entry.get();
        valor_str = self.valor_entry.get().replace(',', '.')
        if not all([data_transacao, descricao, valor_str]): messagebox.showerror("Erro", "Campos obrigatórios.",
                                                                                 parent=self); return
        try:
            valor = float(valor_str);
            assert valor > 0
        except (ValueError, AssertionError):
            messagebox.showerror("Erro", "Valor inválido.", parent=self);
            return
        nova_transacao = {'Data': data_transacao, 'Tipo': self.tipo, 'Descricao': descricao, 'Valor': valor}
        self.callback(nova_transacao);
        messagebox.showinfo("Sucesso", f"{self.tipo} adicionada.", parent=self);
        self.destroy()


# --- CLASSE PRINCIPAL (com a nova função de exportação) ---
class FinanceiroApp:
    def __init__(self, root):
        self.root = root;
        self.root.title("Controle Financeiro Empresarial");
        self.root.geometry("1100x700")
        self.root.state('zoomed')
        tb.Style(theme='superhero')

        self.root.protocol("WM_DELETE_WINDOW", self._on_closing)

        os.makedirs(RELATORIOS_DIR, exist_ok=True)
        self.ano_selecionado = "Todos os Anos";
        self.mes_selecionado = "Todos os Meses"
        self.MESES_NOME = {1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril", 5: "Maio", 6: "Junho", 7: "Julho",
                           8: "Agosto", 9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"}
        self.MESES_NUMERO = {v: k for k, v in self.MESES_NOME.items()}
        self.carregar_dados();
        self.criar_widgets_principais();
        self.atualizar_tudo()

    def _on_closing(self):
        """Exibe uma mensagem de confirmação ao tentar fechar a janela principal."""
        if messagebox.askyesno("Sair", "Tem certeza que deseja fechar o programa?", parent=self.root):
            self.root.destroy()

    def carregar_dados(self):
        con = sqlite3.connect(DB_FILE)
        try:
            self.df_transacoes = pd.read_sql_query("SELECT * FROM transacoes", con)
            if not self.df_transacoes.empty: self.df_transacoes['Data'] = pd.to_datetime(self.df_transacoes['Data'])
        except pd.io.sql.DatabaseError:
            self.df_transacoes = pd.DataFrame(columns=['Data', 'Tipo', 'Descricao', 'Valor'])
        try:
            self.df_dividendos = pd.read_sql_query("SELECT * FROM dividendos", con)
        except pd.io.sql.DatabaseError:
            self.df_dividendos = pd.DataFrame(columns=['Evento', 'Credor', 'Descricao', 'Valor', 'Status'])
        con.close()

    def salvar_dados(self):
        con = sqlite3.connect(DB_FILE)
        df_trans_to_save = self.df_transacoes.copy()
        if 'Data' in df_trans_to_save.columns: df_trans_to_save['Data'] = pd.to_datetime(
            df_trans_to_save['Data']).dt.strftime('%Y-%m-%d %H:%M:%S')
        df_trans_to_save.to_sql('transacoes', con, if_exists='replace', index=False)
        self.df_dividendos.to_sql('dividendos', con, if_exists='replace', index=False)
        con.close()

    # <<< FUNÇÃO DE EXPORTAÇÃO PARA EXCEL TOTALMENTE REFEITA COM "CHARME" >>>
    def exportar_evento_para_excel(self, nome_evento):
        filepath = self._get_safe_filename(nome_evento)
        df_evento = self.df_dividendos[self.df_dividendos['Evento'] == nome_evento].dropna(subset=['Credor'])
        df_relatorio = df_evento[['Credor', 'Descricao', 'Valor', 'Status']].rename(columns={'Descricao': 'Descrição'})

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df_relatorio.to_excel(writer, sheet_name='Relatório de Dívidas', index=False, startrow=2)
            sheet = writer.sheets['Relatório de Dívidas']

            # --- Estilos de Formatação ---
            title_font = Font(size=18, bold=True, color="FFFFFF")
            title_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            total_font = Font(bold=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            pago_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            aguardando_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            alternating_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            currency_format = 'R$ #,##0.00'
            center_align = Alignment(horizontal='center', vertical='center')
            right_align = Alignment(horizontal='right', vertical='center')

            # 1. Título Principal
            sheet.merge_cells('A1:D1')
            title_cell = sheet['A1']
            title_cell.value = f"Relatório de Dívidas - Evento: {nome_evento}"
            title_cell.font = title_font;
            title_cell.fill = title_fill;
            title_cell.alignment = center_align
            sheet.row_dimensions[1].height = 35

            # 2. Cabeçalhos da Tabela
            header_row = 3
            sheet.row_dimensions[header_row].height = 22
            for col_num, column_title in enumerate(df_relatorio.columns, 1):
                cell = sheet.cell(row=header_row, column=col_num)
                cell.font = header_font;
                cell.fill = header_fill;
                cell.border = thin_border

            # 3. Linhas de Dados e Condicionais
            start_data_row = 4
            end_data_row = start_data_row + len(df_relatorio) - 1
            for row_idx, row_data in enumerate(sheet.iter_rows(min_row=start_data_row, max_row=end_data_row),
                                               start=start_data_row):
                sheet.row_dimensions[row_idx].height = 20
                if row_idx % 2 == 1:  # Aplica cor alternada
                    for cell in row_data: cell.fill = alternating_fill

                for cell in row_data:
                    cell.border = thin_border
                    cell.alignment = center_align
                    if cell.column == 3:  # Coluna 'Valor'
                        cell.number_format = currency_format
                        cell.alignment = right_align
                    if cell.column == 4:  # Coluna 'Status'
                        if cell.value == "Pago":
                            cell.fill = pago_fill
                        elif cell.value == "Em Aguardo":
                            cell.fill = aguardando_fill

            # 4. Linha de Total
            total_row_num = end_data_row + 1
            sheet.row_dimensions[total_row_num].height = 22
            sheet[f'B{total_row_num}'] = 'TOTAL';
            sheet[f'B{total_row_num}'].font = total_font;
            sheet[f'B{total_row_num}'].alignment = right_align;
            sheet[f'B{total_row_num}'].border = thin_border
            sheet[f'C{total_row_num}'] = f'=SUM(C{start_data_row}:C{end_data_row})'
            sheet[f'C{total_row_num}'].number_format = currency_format;
            sheet[f'C{total_row_num}'].font = total_font;
            sheet[f'C{total_row_num}'].border = thin_border
            sheet[f'D{total_row_num}'].border = thin_border  # Borda na célula vazia ao lado do total
            sheet[f'A{total_row_num}'].border = thin_border  # Borda na célula vazia ao lado do total

            # 5. Ajuste final de largura das colunas
            for col_num, column_title in enumerate(df_relatorio.columns, 1):
                col_letter = get_column_letter(col_num)
                max_length = len(column_title)
                for cell in sheet[col_letter]:
                    if cell.value: max_length = max(max_length, len(str(cell.value)))
                sheet.column_dimensions[col_letter].width = max_length + 5

    # ... O resto da classe FinanceiroApp permanece o mesmo ...
    def _get_safe_filename(self, nome_evento):
        nome_seguro = re.sub(r'[\\/*?:"<>|]', "", nome_evento);
        return os.path.join(RELATORIOS_DIR, f"Relatorio - {nome_seguro}.xlsx")

    def renomear_relatorio_excel(self, nome_antigo, nome_novo):
        old_path = self._get_safe_filename(nome_antigo);
        new_path = self._get_safe_filename(nome_novo)
        try:
            os.rename(old_path, new_path)
        except FileNotFoundError:
            pass

    def excluir_relatorio_excel(self, nome_evento):
        filepath = self._get_safe_filename(nome_evento)
        try:
            os.remove(filepath)
        except FileNotFoundError:
            pass

    # --- MÉTODO MODIFICADO ---
    def criar_widgets_principais(self):
        main_frame = tb.Frame(self.root, padding=20);
        main_frame.pack(expand=True, fill=BOTH)
        header_frame = tb.Frame(main_frame);
        header_frame.pack(fill=X, pady=(0, 10))
        self.lbl_titulo = tb.Label(header_frame, text="Dashboard Financeiro", font=("Helvetica", 24, "bold"),
                                   bootstyle=PRIMARY);
        self.lbl_titulo.pack(side=LEFT, expand=True, anchor="w")
        filtro_frame = tb.Frame(header_frame);
        filtro_frame.pack(side=RIGHT)
        tb.Label(filtro_frame, text="Filtrar por:").pack(side=LEFT, padx=(0, 10))
        self.combo_ano = tb.Combobox(filtro_frame, bootstyle=PRIMARY, width=12, state="readonly");
        self.combo_ano.pack(side=LEFT, padx=5)
        self.combo_ano.bind("<<ComboboxSelected>>", self.on_ano_selecionado)
        self.combo_mes = tb.Combobox(filtro_frame, bootstyle=PRIMARY, width=12, state="disabled");
        self.combo_mes.pack(side=LEFT, padx=5)
        self.combo_mes.bind("<<ComboboxSelected>>", self.on_mes_selecionado)
        top_frame = tb.Frame(main_frame);
        top_frame.pack(fill=X, pady=10)
        resumo_frame = tb.Labelframe(top_frame, text=" Resumo Financeiro ", bootstyle=INFO);
        resumo_frame.pack(side=LEFT, fill=BOTH, expand=True, padx=10)
        action_frame = tb.Labelframe(top_frame, text=" Ações Rápidas ", bootstyle=INFO);
        action_frame.pack(side=RIGHT, fill=BOTH, expand=True, padx=10)
        self.lbl_receita_total = tb.Label(resumo_frame, text="Receita Total: R$ 0,00", font=("Helvetica", 14),
                                          bootstyle=SUCCESS);
        self.lbl_receita_total.pack(pady=5, anchor="w")
        self.lbl_despesa_total = tb.Label(resumo_frame, text="Despesa Total: R$ 0,00", font=("Helvetica", 14),
                                          bootstyle=DANGER);
        self.lbl_despesa_total.pack(pady=5, anchor="w")
        self.lbl_saldo_total = tb.Label(resumo_frame, text="Saldo Atual: R$ 0,00", font=("Helvetica", 16, "bold"),
                                        bootstyle=PRIMARY);
        self.lbl_saldo_total.pack(pady=10, anchor="w")
        tb.Button(action_frame, text="Adicionar Receita", command=self.abrir_dialogo_receita, bootstyle=SUCCESS).pack(
            pady=5, fill=X)
        tb.Button(action_frame, text="Adicionar Despesa", command=self.abrir_dialogo_despesa, bootstyle=DANGER).pack(
            pady=5, fill=X)
        tb.Button(action_frame, text="Gerenciar Eventos e Dívidas", command=self.abrir_gerenciador_eventos,
                  bootstyle=WARNING).pack(pady=5, fill=X)
        tb.Button(action_frame, text="Gerar Relatório Anual", command=self.abrir_relatorio_mensal, bootstyle=INFO).pack(
            pady=5, fill=X)
        bottom_frame = tb.Frame(main_frame);
        bottom_frame.pack(fill=BOTH, expand=True, pady=10)
        grafico_frame = tb.Labelframe(bottom_frame, text=" Gráfico de Receitas vs. Despesas ", bootstyle=INFO);
        grafico_frame.pack(side=LEFT, fill=BOTH, expand=True, padx=10)
        transacoes_frame = tb.Labelframe(bottom_frame, text=" Últimas Transações ", bootstyle=INFO);
        transacoes_frame.pack(side=RIGHT, fill=BOTH, expand=True, padx=10)
        self.fig, self.ax = plt.subplots(facecolor='#2a3a4a');
        self.canvas = FigureCanvasTkAgg(self.fig, master=grafico_frame);
        self.canvas.get_tk_widget().pack(fill=BOTH, expand=True)

        # Definição das colunas e seus cabeçalhos
        cols = ('Data', 'Tipo', 'Descricao', 'Valor');
        self.tree_transacoes = tb.Treeview(transacoes_frame, columns=cols, show='headings', bootstyle=PRIMARY);
        self.tree_transacoes.pack(fill=BOTH, expand=True)

        # Configuração dos cabeçalhos
        self.tree_transacoes.heading('Data', text='Data')
        self.tree_transacoes.heading('Tipo', text='Tipo')
        self.tree_transacoes.heading('Descricao', text='Categoria/Descrição')
        self.tree_transacoes.heading('Valor', text='Valor')

        # Configuração do alinhamento e largura das colunas
        self.tree_transacoes.column('Data', width=100, anchor=CENTER)
        self.tree_transacoes.column('Tipo', width=80, anchor=CENTER)
        self.tree_transacoes.column('Descricao', width=250, anchor=W)  # W = West (Esquerda)
        self.tree_transacoes.column('Valor', width=120, anchor=E)  # E = East (Direita)

        tb.Button(transacoes_frame, text="Excluir Transação Selecionada", command=self.excluir_transacao,
                  bootstyle=(DANGER, OUTLINE)).pack(pady=5)

    def _on_transacao_adicionada(self, nova_transacao):
        nova_transacao['Data'] = pd.to_datetime(nova_transacao['Data'])
        self.df_transacoes = pd.concat([self.df_transacoes, pd.DataFrame([nova_transacao])], ignore_index=True)
        self.salvar_dados();
        self.atualizar_tudo()

    def registrar_pagamento_divida(self, nome_evento, credor, descricao_divida, valor_divida):
        """Cria uma nova transação 'Despesa' quando uma dívida é marcada como paga."""
        data_pagamento = pd.to_datetime('today').strftime('%Y-%m-%d')
        descricao_transacao = f"Pagamento p/ {credor} ({nome_evento}): {descricao_divida}"

        nova_transacao = {
            'Data': data_pagamento,
            'Tipo': 'Despesa',
            'Descricao': descricao_transacao,
            'Valor': valor_divida
        }
        self._on_transacao_adicionada(nova_transacao)
        messagebox.showinfo(
            "Despesa Registrada",
            f"Uma nova despesa de R$ {valor_divida:,.2f} foi registrada no seu fluxo de caixa.",
            parent=self.root
        )

    def atualizar_seletores(self):
        if not self.df_transacoes.empty:
            anos = sorted(self.df_transacoes['Data'].dt.year.unique(), reverse=True)
            self.combo_ano['values'] = ["Todos os Anos"] + anos
        else:
            self.combo_ano['values'] = ["Todos os Anos"]
        self.combo_ano.set(self.ano_selecionado)
        if self.ano_selecionado != "Todos os Anos":
            ano = int(self.ano_selecionado)
            meses_ano = self.df_transacoes[self.df_transacoes['Data'].dt.year == ano]
            if not meses_ano.empty:
                meses_num = sorted(meses_ano['Data'].dt.month.unique())
                self.combo_mes['values'] = ["Todos os Meses"] + [self.MESES_NOME[m] for m in meses_num]
                self.combo_mes.config(state="readonly")
            else:
                self.combo_mes['values'] = ["Todos os Meses"];
                self.combo_mes.config(state="disabled")
        else:
            self.combo_mes['values'] = [];
            self.combo_mes.config(state="disabled")
        self.combo_mes.set(self.mes_selecionado)

    def on_ano_selecionado(self, event=None):
        self.ano_selecionado = self.combo_ano.get()
        self.mes_selecionado = "Todos os Meses";
        self.atualizar_tudo()

    def on_mes_selecionado(self, event=None):
        self.mes_selecionado = self.combo_mes.get();
        self.atualizar_tudo()

    def abrir_dialogo_receita(self):
        TransacaoDialog(self.root, 'Receita', self._on_transacao_adicionada)

    def abrir_dialogo_despesa(self):
        TransacaoDialog(self.root, 'Despesa', self._on_transacao_adicionada)

    def abrir_gerenciador_eventos(self):
        EventosManagerWindow(self.root, self)

    def abrir_relatorio_mensal(self):
        RelatorioMensalWindow(self.root, self)

    def excluir_transacao(self):
        selecionado = self.tree_transacoes.selection()
        if not selecionado or not messagebox.askyesno("Confirmar", "Tem certeza?"): return
        indices_para_excluir = [int(item) for item in selecionado]
        self.df_transacoes = self.df_transacoes.drop(indices_para_excluir).reset_index(drop=True)
        self.salvar_dados();
        self.atualizar_tudo();
        messagebox.showinfo("Sucesso", "Transação excluída.")

    def atualizar_tudo(self):
        df_filtrado = self.df_transacoes.copy()
        if self.ano_selecionado == "Todos os Anos":
            titulo = "Dashboard Financeiro - Visão Geral"
        else:
            ano = int(self.ano_selecionado)
            df_filtrado = df_filtrado[df_filtrado['Data'].dt.year == ano]
            titulo = f"Dashboard Financeiro - {ano}"
            if self.mes_selecionado != "Todos os Meses":
                mes = self.MESES_NUMERO[self.mes_selecionado]
                df_filtrado = df_filtrado[df_filtrado['Data'].dt.month == mes]
                titulo += f" / {self.mes_selecionado}"
            else:
                titulo += " (Visão Anual)"
        self.lbl_titulo.config(text=titulo)
        self.atualizar_seletores()
        self.atualizar_resumo(df_filtrado)
        self.atualizar_lista_transacoes(df_filtrado)
        self.atualizar_grafico(df_filtrado)

    def atualizar_resumo(self, df_filtrado):
        receitas = df_filtrado[df_filtrado['Tipo'] == 'Receita']['Valor'].sum()
        despesas = df_filtrado[df_filtrado['Tipo'] == 'Despesa']['Valor'].sum()
        self.lbl_receita_total.config(text=f"Receita Total: R$ {receitas:,.2f}");
        self.lbl_despesa_total.config(text=f"Despesa Total: R$ {despesas:,.2f}");
        self.lbl_saldo_total.config(text=f"Saldo Atual: R$ {receitas - despesas:,.2f}")

    # --- MÉTODO MODIFICADO ---
    def atualizar_lista_transacoes(self, df_filtrado):
        self.tree_transacoes.delete(*self.tree_transacoes.get_children())
        df_sorted = df_filtrado.sort_values(by='Data', ascending=False)
        for index, row in df_sorted.head(20).iterrows():
            # Formata os valores para exibição
            data_str = row['Data'].strftime('%Y-%m-%d')
            tipo_str = row['Tipo']
            desc_str = row['Descricao']
            valor_str = f"R$ {row['Valor']:,.2f}"  # Formata o valor como moeda

            # Insere os valores formatados na Treeview
            self.tree_transacoes.insert('', 'end', iid=index, values=(data_str, tipo_str, desc_str, valor_str))

    def atualizar_grafico(self, df_filtrado):
        self.ax.clear()
        total_receitas = df_filtrado[df_filtrado['Tipo'] == 'Receita']['Valor'].sum()
        total_despesas = df_filtrado[df_filtrado['Tipo'] == 'Despesa']['Valor'].sum()
        labels = ['Receitas', 'Despesas'];
        valores = [total_receitas, total_despesas];
        cores = ['#28a745', '#dc3545']
        bars = self.ax.bar(labels, valores, color=cores)
        self.ax.set_title("Visão Geral do Período", color='white');
        self.ax.set_ylabel("Valor (R$)", color='white')
        self.ax.tick_params(axis='x', colors='white');
        self.ax.tick_params(axis='y', colors='white')
        self.ax.set_facecolor('#2a3a4a')
        formatter = plt.FuncFormatter(lambda x, p: f'R$ {x:,.0f}');
        self.ax.yaxis.set_major_formatter(formatter)
        for bar in bars:
            yval = bar.get_height()
            self.ax.text(bar.get_x() + bar.get_width() / 2.0, yval, f'R$ {yval:,.2f}', va='bottom', ha='center',
                         color='white')
        self.fig.tight_layout();
        self.canvas.draw()


if __name__ == "__main__":
    root = tb.Window(themename="superhero")
    app = FinanceiroApp(root)
    root.mainloop()